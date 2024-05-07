import openpyxl

# Create a new workbook object
wb = openpyxl.Workbook()

# Get the active sheet from the workbook
sheet = wb.active  # Default sheet is named 'Sheet', which is the active one when a new workbook is created

# SHEETS
sheet.title = 'example' # We can change the name

# wb.get_sheet_names()   -> Create sheet with name Sheet by default
new_sheet = wb.create_sheet('New') # We can create new sheet
my_new_sheet = wb.create_sheet(index=0, title='My Other Sheet')

# CELLS
sheet['A1'] = 'Data'             # Create data in cell A1, Sheet name
new_sheet['A1'] = 'Time'         # Create data in cell A1, New_Sheet name
my_new_sheet['A1'] = 'Minutes'   # Create data in cell A1, My_new_sheet

# Print the value of cell A1
print("Value in cell A1:", sheet)

# Print the value of cell B1
print("Value in cell B1:", new_sheet)

# You can print all sheet names if needed
print("All sheet names in the workbook:", wb.sheetnames)

# Don't forget to save the workbook if you have made changes
wb.save('example.xlsx')


# Create new sheet or use existing
# wb.create_sheet('New Sheet Name') or wb['SheetName']