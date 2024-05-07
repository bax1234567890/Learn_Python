import openpyxl
import os

# Change Python to working directory
os.chdir('/Users/ionbota/Documents')

# Navigate to excel
workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)
sheet = workbook['Sheet1']
sheet_names = workbook.sheetnames
for name in sheet_names:
    print(name)
type(sheet)

# In particula cell extract information
cellA1 = sheet['A1']
valueA1 = cellA1.value
print(str(valueA1))

cellB1 = sheet['B1']
valueB1 = cellB1.value
print(str(valueB1))

cellC1 = sheet['C1']
valueC1 = cellC1.value
print(str(valueC1))

newcell = sheet.cell(row=1, column=2)
print(str(newcell))

# use For to display all values
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
